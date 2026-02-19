"""
Scan datasets in bench/tabpfn_paper/data, auto-detect and load them,
then output a comprehensive statistical ID card for each (console + JSON + PNG + PDF).

By default, only datasets listed in data/DatabaseDetail.xlsx (Database/Dataset columns)
are scanned. Use --all to scan every detected dataset.

Usage:
    cd nirs4all/
    python bench/tabpfn_paper/scan_datasets.py                     # selected datasets only
    python bench/tabpfn_paper/scan_datasets.py --all               # all datasets
    python bench/tabpfn_paper/scan_datasets.py --json results.json
    python bench/tabpfn_paper/scan_datasets.py --category regression
    python bench/tabpfn_paper/scan_datasets.py --no-plots          # skip figure generation
    python bench/tabpfn_paper/scan_datasets.py --from-json scan_results.json  # regenerate figures from existing JSON
"""

import argparse
import json
import sys
import traceback
from pathlib import Path

import matplotlib
import numpy as np
from scipy.stats import kurtosis as scipy_kurtosis

matplotlib.use("Agg")
import contextlib

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.gridspec import GridSpec

from nirs4all.data.config import DatasetConfigs
from nirs4all.data.parsers.folder_parser import FolderParser

DATA_ROOT = Path(__file__).parent / "data"
SELECTION_XLSX = DATA_ROOT / "DatabaseDetail.xlsx"

# ── Matplotlib style ──────────────────────────────────────────────────
COLORS = {
    "bg": "#FAFAFA",
    "panel": "#FFFFFF",
    "text": "#1A1A2E",
    "muted": "#6B7280",
    "accent": "#0F766E",       # teal-700
    "band_inner": "#99F6E4",   # teal-200
    "band_outer": "#CCFBF1",   # teal-100
    "mean_line": "#0D9488",    # teal-600
    "hist": "#14B8A6",         # teal-500
    "hist_edge": "#0F766E",
    "grid": "#E5E7EB",
    "divider": "#D1D5DB",
    "error_bg": "#FEF2F2",
    "error_text": "#B91C1C",
}

def load_selected_suffixes(xlsx_path: Path) -> set[str]:
    """Load selected Database/Dataset pairs from the Excel file.

    Returns a set of normalised path suffixes like "ALPINE/ALPINE_C_424_KS".
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return set()

    header = [str(c).strip() if c else "" for c in rows[0]]
    db_idx = header.index("Database")
    ds_idx = header.index("Dataset")

    suffixes = set()
    for row in rows[1:]:
        db = str(row[db_idx]).strip() if row[db_idx] else ""
        ds = str(row[ds_idx]).strip() if row[ds_idx] else ""
        if db and ds:
            suffixes.add(f"{db}/{ds}")
    print(suffixes)
    return suffixes

def find_datasets(root: Path, selected_suffixes: set[str] | None = None) -> list[dict]:
    """Recursively find dataset folders using FolderParser (same logic as webapp scan).

    When *selected_suffixes* is given, only datasets whose path ends with one of
    the ``Database/Dataset`` suffixes from the Excel selection file are kept.
    """
    parser = FolderParser()
    datasets = []
    _recurse(root, parser, datasets, [])
    if selected_suffixes is not None:
        datasets = [d for d in datasets if _matches_selection(d["path"], selected_suffixes)]
    return datasets

def _matches_selection(dataset_path: str, suffixes: set[str]) -> bool:
    """Return True if *dataset_path* ends with one of the selected Database/Dataset suffixes."""
    parts = Path(dataset_path).parts
    if len(parts) < 2:
        return False
    tail = f"{parts[-2]}/{parts[-1]}"
    return tail in suffixes

def _recurse(folder: Path, parser: FolderParser, results: list, groups: list[str]):
    """Recurse into folder tree. If a folder is a valid dataset, stop recursion there."""
    if not folder.is_dir():
        return

    result = parser.parse(str(folder))
    if result.success:
        results.append({
            "path": str(folder),
            "name": result.dataset_name,
            "config": result.config,
            "groups": list(groups),
            "warnings": result.warnings,
        })
        return  # Don't recurse into dataset folders

    # Not a dataset — recurse into children
    for child in sorted(folder.iterdir()):
        if child.is_dir() and not child.name.startswith('.'):
            _recurse(child, parser, results, groups + [folder.name])

def build_id_card(info: dict) -> dict:
    """Load dataset and compute statistical ID card."""
    config = info["config"]
    name = info["name"]

    # Allow NAs instead of aborting — we just want statistics
    if config.get("global_params") is None:
        config["global_params"] = {}
    config["global_params"]["na_policy"] = "ignore"

    # Strip empty metadata files that would cause load errors
    for key in ["train_group", "test_group"]:
        path = config.get(key)
        if path and Path(path).exists() and Path(path).stat().st_size == 0:
            config[key] = None

    configs = DatasetConfigs(config)
    datasets = configs.get_datasets()
    if not datasets:
        raise ValueError("No dataset could be loaded from config")

    ds = datasets[0]

    # Gather all samples (train + test combined)
    partitions = []
    x_parts = []
    y_parts = []
    for part in ["train", "test"]:
        try:
            x = ds.x({"partition": part})
            if x is not None and len(x) > 0:
                x_parts.append(x)
                partitions.append(part)
        except Exception:
            pass
        try:
            y = ds.y({"partition": part})
            if y is not None and len(y) > 0:
                y_parts.append(y)
        except Exception:
            pass

    x_all = np.concatenate(x_parts, axis=0) if x_parts else np.empty((0, 0))
    y_all = np.concatenate(y_parts, axis=0) if y_parts else np.empty((0,))

    # Flatten y if multi-dim
    if y_all.ndim > 1 and y_all.shape[1] == 1:
        y_all = y_all.ravel()

    n_samples = x_all.shape[0]
    n_features = x_all.shape[1] if x_all.ndim > 1 else 0

    # Train/test split sizes
    n_train = x_parts[0].shape[0] if len(x_parts) > 0 and partitions[0] == "train" else 0
    n_test = 0
    if len(x_parts) > 1:
        n_test = x_parts[1].shape[0]
    elif len(x_parts) == 1 and partitions[0] == "test":
        n_test = x_parts[0].shape[0]
        n_train = 0

    # Target name — from Y file header
    target_name = _extract_target_name(config)

    # Task type — force from folder path when available, otherwise use detected
    category = _detect_category(info["path"])
    if category == "regression":
        ds.set_task_type("regression", forced=True)
    elif category == "classification" and not ds.is_classification:
        ds.set_task_type("multiclass_classification", forced=True)
    task_type = ds.task_type.value if ds.task_type else "unknown"

    # Signal type — use library detection, then heuristic fallback from X value range
    try:
        signal_type = ds.signal_type(0).value
    except Exception:
        signal_type = "unknown"
    if signal_type in ("unknown", "auto") and n_samples > 0 and n_features > 0:
        x_min, x_max = float(np.nanmin(x_all)), float(np.nanmax(x_all))
        if x_min >= -0.5 and x_max <= 4.0:
            signal_type = "absorbance"
        elif (x_min >= 0 and x_max <= 1.1) or (x_min >= 0 and x_max <= 110):
            signal_type = "reflectance" if x_max <= 1.1 else "reflectance%"

    # Header unit
    try:
        header_unit = ds.header_unit(0)
    except Exception:
        header_unit = "unknown"

    # Metadata — detect repetition columns among ID candidates
    metadata_names = ds.metadata_columns if ds.metadata_columns else []
    id_candidates = [c for c in metadata_names if "id" in c.lower()]
    name_repetitions = [c for c in metadata_names if "repetition" in c.lower() or "repeat" in c.lower()]

    # A column is a repetition col if: grouping by it, every group has constant y,
    # and at least some groups have >1 sample (i.e. actual repetitions exist).
    repetition_columns = list(name_repetitions)  # start with name-based matches
    for col in id_candidates:
        if col in repetition_columns:
            continue
        if _is_repetition_column(ds, col, y_all, partitions):
            repetition_columns.append(col)

    # Feature statistics
    feat_stats = {}
    if n_features > 0 and n_samples > 0:
        feat_stats = {
            "min": float(np.nanmin(x_all)),
            "max": float(np.nanmax(x_all)),
            "mean": float(np.nanmean(x_all)),
            "std": float(np.nanstd(x_all)),
        }

    # Target statistics
    target_stats = {}
    if len(y_all) > 0:
        y_numeric = y_all[~np.isnan(y_all.astype(float))] if np.issubdtype(y_all.dtype, np.floating) else y_all
        if len(y_numeric) > 0:
            target_stats["min"] = float(np.nanmin(y_numeric))
            target_stats["max"] = float(np.nanmax(y_numeric))
            target_stats["mean"] = float(np.nanmean(y_numeric))
            target_stats["std"] = float(np.nanstd(y_numeric))
            try:
                target_stats["kurtosis"] = float(scipy_kurtosis(y_numeric.astype(float), fisher=True, nan_policy='omit'))
            except Exception:
                target_stats["kurtosis"] = None
            if ds.is_classification:
                unique, counts = np.unique(y_all, return_counts=True)
                target_stats["classes"] = {str(u): int(c) for u, c in zip(unique, counts, strict=False)}
                target_stats["n_classes"] = len(unique)

    # Wavelength range + axis values for plotting
    wavelength_range = None
    wavelengths = None
    try:
        headers = ds.headers(0)
        if headers:
            numeric_headers = []
            for h in headers:
                with contextlib.suppress(ValueError, TypeError):
                    numeric_headers.append(float(h))
            if len(numeric_headers) > 1:
                wavelengths = numeric_headers
                wavelength_range = {
                    "min": min(numeric_headers),
                    "max": max(numeric_headers),
                    "unit": header_unit,
                }
    except Exception:
        pass

    # Spectra percentile profiles for plotting
    plot_data = {}
    if n_features > 0 and n_samples > 0:
        plot_data["wavelengths"] = wavelengths or list(range(n_features))
        plot_data["spectra_mean"] = np.nanmean(x_all, axis=0).tolist()
        plot_data["spectra_p5"] = np.nanpercentile(x_all, 5, axis=0).tolist()
        plot_data["spectra_p25"] = np.nanpercentile(x_all, 25, axis=0).tolist()
        plot_data["spectra_p75"] = np.nanpercentile(x_all, 75, axis=0).tolist()
        plot_data["spectra_p95"] = np.nanpercentile(x_all, 95, axis=0).tolist()
    if len(y_all) > 0:
        plot_data["y_values"] = y_all.ravel().tolist()

    card = {
        "name": name,
        "path": info["path"],
        "groups": info["groups"],
        "category": _detect_category(info["path"]),
        "task_type": task_type,
        "signal_type": signal_type,
        "header_unit": header_unit,
        "n_samples": n_samples,
        "n_train": n_train,
        "n_test": n_test,
        "n_features": n_features,
        "target_name": target_name,
        "metadata_columns": metadata_names,
        "repetition_columns": repetition_columns,
        "wavelength_range": wavelength_range,
        "features": feat_stats,
        "targets": target_stats,
        "warnings": info.get("warnings", []),
        "plot_data": plot_data,
    }
    return card

def _is_repetition_column(ds, col: str, y_all: np.ndarray, partitions: list[str]) -> bool:
    """Check if a metadata column identifies sample repetitions.

    A repetition column means: all samples sharing the same column value also
    share the same target y.  At least some groups must have >1 sample.
    """
    try:
        # Collect column values aligned with y across partitions
        col_parts = []
        for part in partitions:
            vals = ds.metadata_column(col, {"partition": part})
            if vals is not None and len(vals) > 0:
                col_parts.append(vals)
        if not col_parts:
            return False
        col_all = np.concatenate(col_parts)
    except Exception:
        return False

    if len(col_all) != len(y_all) or len(y_all) == 0:
        return False

    # Group by column value, check y is constant within each group
    y_flat = y_all.ravel().astype(float)
    groups: dict[str, list[float]] = {}
    for cval, yval in zip(col_all, y_flat, strict=False):
        key = str(cval)
        groups.setdefault(key, []).append(yval)

    has_repeated = False
    for members in groups.values():
        if len(members) > 1:
            has_repeated = True
            # Constant y within group? Use tolerance for float comparison
            arr = np.array(members)
            if np.nanstd(arr) > 1e-8 * (abs(np.nanmean(arr)) + 1e-12):
                return False  # y differs within a group → not a repetition col

    return has_repeated  # True only if at least one group had >1 sample

def _extract_target_name(config: dict) -> str:
    """Extract target column name from Y file header."""
    y_path = config.get("train_y") or config.get("test_y")
    if not y_path:
        return "unknown"
    try:
        p = Path(y_path)
        if not p.exists():
            return "unknown"
        with open(p, encoding="utf-8") as f:
            first_line = f.readline().strip()
        # If the first line is not numeric, it's a header
        try:
            float(first_line.replace(",", ".").replace(";", ""))
            return "unknown"  # No header, just numbers
        except ValueError:
            # Remove delimiter if present
            for delim in [";", ",", "\t"]:
                if delim in first_line:
                    parts = [p.strip() for p in first_line.split(delim) if p.strip()]
                    return parts[0] if len(parts) == 1 else ", ".join(parts)
            return first_line
    except Exception:
        return "unknown"

def _detect_category(path: str) -> str:
    """Detect if dataset is regression or classification from path."""
    path_lower = path.lower().replace("\\", "/")
    if "/classification/" in path_lower:
        return "classification"
    if "/regression/" in path_lower:
        return "regression"
    return "unknown"

# ── Figure rendering ──────────────────────────────────────────────────

def _fmt(v, precision=4):
    """Format a number nicely for display."""
    if v is None:
        return "n/a"
    if isinstance(v, int) or (isinstance(v, float) and v == int(v) and abs(v) < 1e6):
        return str(int(v))
    return f"{v:.{precision}g}"

def render_card_figure(card: dict, output_path: Path):
    """Render a single dataset ID card as a PNG figure."""
    pd = card.get("plot_data", {})
    has_spectra = "spectra_mean" in pd
    has_y = "y_values" in pd

    fig = plt.figure(figsize=(16, 8.5), facecolor=COLORS["bg"], dpi=150)
    gs = GridSpec(2, 2, figure=fig, height_ratios=[1, 0.55],
                 hspace=0.28, wspace=0.25,
                 left=0.06, right=0.97, top=0.87, bottom=0.06)

    # ── Title ─────────────────────────────────────────────────────────
    title = card["name"]
    category_badge = card["category"].upper() if card["category"] != "unknown" else ""
    task_badge = card["task_type"].replace("_", " ").upper()
    fig.suptitle(title, fontsize=16, fontweight="bold", color=COLORS["text"],
                 x=0.06, ha="left", y=0.97)
    badge_text = f"{category_badge}  |  {task_badge}  |  {card['n_samples']} samples  |  {card['n_features']} features"
    fig.text(0.06, 0.935, badge_text, fontsize=9, color=COLORS["muted"],
             fontfamily="monospace")

    # ── Top-left: Spectra chart ───────────────────────────────────────
    ax_spectra = fig.add_subplot(gs[0, 0])
    if has_spectra:
        wl = np.array(pd["wavelengths"])
        mean = np.array(pd["spectra_mean"])
        p5 = np.array(pd["spectra_p5"])
        p25 = np.array(pd["spectra_p25"])
        p75 = np.array(pd["spectra_p75"])
        p95 = np.array(pd["spectra_p95"])

        ax_spectra.fill_between(wl, p5, p95, alpha=0.35, color=COLORS["band_outer"],
                                label="5%-95%", linewidth=0)
        ax_spectra.fill_between(wl, p25, p75, alpha=0.50, color=COLORS["band_inner"],
                                label="25%-75%", linewidth=0)
        ax_spectra.plot(wl, mean, color=COLORS["mean_line"], linewidth=1.3, label="Mean")

        unit = card.get("header_unit", "")
        xlabel = f"Wavelength ({unit})" if unit and unit not in ("unknown", "none", "text", "index") else "Feature index"
        ax_spectra.set_xlabel(xlabel, fontsize=9, color=COLORS["muted"])
        sig = card.get("signal_type", "")
        ylabel = sig.capitalize() if sig and sig != "unknown" else "Intensity"
        ax_spectra.set_ylabel(ylabel, fontsize=9, color=COLORS["muted"])
        ax_spectra.legend(fontsize=7, loc="upper right", framealpha=0.8)

        # Feature stats annotation
        feat = card.get("features", {})
        if feat:
            stats_txt = (f"min={_fmt(feat['min'])}  max={_fmt(feat['max'])}\n"
                         f"mean={_fmt(feat['mean'])}  std={_fmt(feat['std'])}")
            ax_spectra.text(0.02, 0.97, stats_txt, transform=ax_spectra.transAxes,
                            fontsize=7, fontfamily="monospace", verticalalignment="top",
                            color=COLORS["muted"],
                            bbox={"boxstyle": "round,pad=0.3", "fc": "white", "ec": COLORS["grid"], "alpha": 0.9})
    else:
        ax_spectra.text(0.5, 0.5, "No spectra data", ha="center", va="center",
                        fontsize=11, color=COLORS["muted"])
        ax_spectra.set_xticks([])
        ax_spectra.set_yticks([])

    ax_spectra.set_title("Spectra", fontsize=10, fontweight="bold",
                         color=COLORS["text"], loc="left")
    _style_axis(ax_spectra)

    # ── Top-right: Y histogram ────────────────────────────────────────
    ax_hist = fig.add_subplot(gs[0, 1])
    if has_y:
        y_vals = np.array(pd["y_values"], dtype=float)
        y_clean = y_vals[np.isfinite(y_vals)]
        is_classif = card["task_type"] in ("binary_classification", "multiclass_classification")

        if is_classif:
            unique, counts = np.unique(y_clean, return_counts=True)
            labels = [str(int(u)) if u == int(u) else str(u) for u in unique]
            bars = ax_hist.bar(labels, counts, color=COLORS["hist"],
                               edgecolor=COLORS["hist_edge"], linewidth=0.5)
            ax_hist.set_xlabel("Class", fontsize=9, color=COLORS["muted"])
            ax_hist.set_ylabel("Count", fontsize=9, color=COLORS["muted"])
            # Annotate counts on bars
            for bar, c in zip(bars, counts, strict=False):
                ax_hist.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                             str(c), ha="center", va="bottom", fontsize=7, color=COLORS["muted"])
        else:
            n_bins = min(50, max(15, len(y_clean) // 8))
            ax_hist.hist(y_clean, bins=n_bins, color=COLORS["hist"],
                         edgecolor=COLORS["hist_edge"], linewidth=0.5, alpha=0.85)
            # Vertical lines for mean ± std
            tgt = card.get("targets", {})
            if tgt:
                mu = tgt["mean"]
                sd = tgt["std"]
                ax_hist.axvline(mu, color=COLORS["accent"], linewidth=1.2,
                                linestyle="-", label=f"mean={_fmt(mu)}")
                ax_hist.axvline(mu - sd, color=COLORS["accent"], linewidth=0.8,
                                linestyle="--", alpha=0.6, label=f"std={_fmt(sd)}")
                ax_hist.axvline(mu + sd, color=COLORS["accent"], linewidth=0.8,
                                linestyle="--", alpha=0.6)
                ax_hist.legend(fontsize=7, loc="upper right", framealpha=0.8)
            ax_hist.set_xlabel(card.get("target_name", "Y"), fontsize=9, color=COLORS["muted"])
            ax_hist.set_ylabel("Count", fontsize=9, color=COLORS["muted"])

        # Target stats annotation
        tgt = card.get("targets", {})
        if tgt:
            lines = [f"min={_fmt(tgt.get('min'))}  max={_fmt(tgt.get('max'))}",
                     f"mean={_fmt(tgt.get('mean'))}  std={_fmt(tgt.get('std'))}"]
            if tgt.get("kurtosis") is not None:
                lines.append(f"kurtosis={_fmt(tgt['kurtosis'])}")
            if "n_classes" in tgt:
                lines.append(f"classes={tgt['n_classes']}")
            stats_txt = "\n".join(lines)
            ax_hist.text(0.02, 0.97, stats_txt, transform=ax_hist.transAxes,
                         fontsize=7, fontfamily="monospace", verticalalignment="top",
                         color=COLORS["muted"],
                         bbox={"boxstyle": "round,pad=0.3", "fc": "white", "ec": COLORS["grid"], "alpha": 0.9})
    else:
        ax_hist.text(0.5, 0.5, "No target data", ha="center", va="center",
                     fontsize=11, color=COLORS["muted"])
        ax_hist.set_xticks([])
        ax_hist.set_yticks([])

    hist_title = f"Target: {card.get('target_name', 'Y')}"
    ax_hist.set_title(hist_title, fontsize=10, fontweight="bold",
                      color=COLORS["text"], loc="left")
    _style_axis(ax_hist)

    # ── Bottom: Info panel ────────────────────────────────────────────
    ax_info = fig.add_subplot(gs[1, :])
    ax_info.set_xlim(0, 1)
    ax_info.set_ylim(0, 1)
    ax_info.axis("off")
    ax_info.set_facecolor(COLORS["panel"])
    # Draw a top border
    ax_info.axhline(y=1.0, color=COLORS["divider"], linewidth=1.0, xmin=0, xmax=1)

    info_lines = _build_info_lines(card)
    n_cols = 2
    col_items = [[], []]
    for i, line in enumerate(info_lines):
        col_items[i % n_cols].append(line)

    for col_idx, items in enumerate(col_items):
        x = 0.02 + col_idx * 0.50
        for row_idx, (label, value) in enumerate(items):
            y = 0.88 - row_idx * 0.18
            ax_info.text(x, y, label, fontsize=8, fontweight="bold",
                         color=COLORS["text"], verticalalignment="top",
                         fontfamily="monospace", transform=ax_info.transAxes)
            ax_info.text(x + 0.12, y, value, fontsize=8,
                         color=COLORS["muted"], verticalalignment="top",
                         fontfamily="monospace", transform=ax_info.transAxes)

    fig.savefig(output_path, dpi=150, facecolor=COLORS["bg"])
    plt.close(fig)

def _build_info_lines(card: dict) -> list[tuple[str, str]]:
    """Build label-value pairs for the info panel."""
    lines = []
    lines.append(("Samples", f"{card['n_samples']}  (train: {card['n_train']}, test: {card['n_test']})"))
    lines.append(("Features", str(card["n_features"])))

    wr = card.get("wavelength_range")
    if wr:
        lines.append(("Wavelengths", f"{_fmt(wr['min'])} - {_fmt(wr['max'])} {wr['unit']}"))
    else:
        lines.append(("Wavelengths", "(none)"))

    lines.append(("Signal type", card.get("signal_type", "unknown")))

    meta = card.get("metadata_columns", [])
    if meta:
        meta_str = ", ".join(meta[:6])
        if len(meta) > 6:
            meta_str += f"  (+{len(meta) - 6} more)"
        lines.append(("Metadata", meta_str))
    else:
        lines.append(("Metadata", "(none)"))

    rep_cols = card.get("repetition_columns", [])
    if rep_cols:
        lines.append(("Repetitions", ", ".join(rep_cols)))

    path = card.get("path", "")
    parts = Path(path).parts
    path = str(Path(*parts[-2:])) if len(parts) >= 2 else path
    lines.append(("Path", path))

    return lines

def _style_axis(ax):
    """Apply consistent axis styling."""
    ax.set_facecolor(COLORS["panel"])
    ax.tick_params(labelsize=7, colors=COLORS["muted"])
    ax.grid(True, alpha=0.4, color=COLORS["grid"], linewidth=0.5)
    for spine in ax.spines.values():
        spine.set_color(COLORS["divider"])
        spine.set_linewidth(0.5)

def render_error_figure(error: dict, output_path: Path):
    """Render a single error page as a PNG figure."""
    fig, ax = plt.subplots(figsize=(16, 8.5), facecolor=COLORS["error_bg"], dpi=150)
    ax.axis("off")

    fig.suptitle(f"LOAD ERROR: {error['name']}", fontsize=16, fontweight="bold",
                 color=COLORS["error_text"], x=0.06, ha="left", y=0.92)
    fig.text(0.06, 0.85, error["path"], fontsize=9, color=COLORS["muted"],
             fontfamily="monospace")

    error_msg = error["error"]
    # Wrap long error messages
    wrapped = "\n".join(error_msg[i:i+120] for i in range(0, len(error_msg), 120))
    ax.text(0.03, 0.70, wrapped, transform=ax.transAxes,
            fontsize=10, fontfamily="monospace", color=COLORS["error_text"],
            verticalalignment="top")

    fig.savefig(output_path, dpi=150, facecolor=COLORS["error_bg"])
    plt.close(fig)

def render_pdf(cards: list[dict], errors: list[dict], pdf_path: Path, png_dir: Path):
    """Combine all card PNGs + error PNGs into a single PDF."""
    with PdfPages(str(pdf_path)) as pdf:
        for card in cards:
            png = png_dir / f"{card['name']}.png"
            if png.exists():
                img = plt.imread(str(png))
                fig, ax = plt.subplots(figsize=(16, 8.5), dpi=150)
                ax.imshow(img)
                ax.axis("off")
                fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
                pdf.savefig(fig, dpi=150)
                plt.close(fig)

        for err in errors:
            png = png_dir / f"{err['name']}_error.png"
            if png.exists():
                img = plt.imread(str(png))
                fig, ax = plt.subplots(figsize=(16, 8.5), dpi=150)
                ax.imshow(img)
                ax.axis("off")
                fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
                pdf.savefig(fig, dpi=150)
                plt.close(fig)

def print_card(card: dict):
    """Print a nicely formatted ID card."""
    w = 72
    print("=" * w)
    print(f"  {card['name']}")
    print(f"  {card['path']}")
    if card["groups"]:
        print(f"  Groups: {' > '.join(card['groups'])}")
    print("-" * w)
    print(f"  Category     : {card['category']}")
    print(f"  Task type    : {card['task_type']}")
    print(f"  Signal type  : {card['signal_type']}")
    print(f"  Header unit  : {card['header_unit']}")
    print(f"  Samples      : {card['n_samples']}  (train: {card['n_train']}, test: {card['n_test']})")
    print(f"  Features     : {card['n_features']}")

    if card["wavelength_range"]:
        wr = card["wavelength_range"]
        print(f"  Wavelengths  : {wr['min']} - {wr['max']} {wr['unit']}")

    print(f"  Target name  : {card['target_name']}")

    if card["metadata_columns"]:
        print(f"  Metadata     : {', '.join(card['metadata_columns'])}")
        if card["repetition_columns"]:
            print(f"    Repetitions: {', '.join(card['repetition_columns'])}")
    else:
        print("  Metadata     : (none)")

    feat = card["features"]
    if feat:
        print(f"  Features X   : min={feat['min']:.4g}  max={feat['max']:.4g}  mean={feat['mean']:.4g}  std={feat['std']:.4g}")

    tgt = card["targets"]
    if tgt:
        print(f"  Targets  Y   : min={tgt['min']:.4g}  max={tgt['max']:.4g}  mean={tgt['mean']:.4g}  std={tgt['std']:.4g}")
        if tgt.get("kurtosis") is not None:
            print(f"               : kurtosis={tgt['kurtosis']:.4g}")
        if "n_classes" in tgt:
            print(f"  Classes      : {tgt['n_classes']}  {tgt['classes']}")

    if card["warnings"]:
        for w_msg in card["warnings"]:
            print(f"  WARNING: {w_msg}")

    print("=" * 72)

def print_error(info: dict, error: str):
    """Print a dataset that failed to load."""
    print("!" * 72)
    print(f"  ERROR: {info['name']}")
    print(f"  Path  : {info['path']}")
    print(f"  {error}")
    print("!" * 72)

def main():
    parser = argparse.ArgumentParser(description="Scan and profile all NIRS datasets")
    parser.add_argument("--json", type=str, default=None, help="Output JSON file path")
    parser.add_argument("--category", type=str, default=None, choices=["regression", "classification"],
                        help="Scan only this category")
    parser.add_argument("--no-plots", action="store_true", help="Skip PNG/PDF generation")
    parser.add_argument("--from-json", type=str, default=None,
                        help="Regenerate figures from an existing JSON file (skip scanning)")
    parser.add_argument("--all", action="store_true",
                        help="Scan all datasets, ignoring DatabaseDetail.xlsx selection")
    args = parser.parse_args()

    out_dir = Path(__file__).parent
    json_path = Path(args.json) if args.json else out_dir / "scan_results.json"
    png_dir = out_dir / "cards"
    pdf_path = out_dir / "scan_cards.pdf"

    # ── Mode: regenerate from JSON ────────────────────────────────────
    if args.from_json:
        src = Path(args.from_json)
        if not src.exists():
            print(f"JSON file not found: {src}")
            sys.exit(1)
        with open(src, encoding="utf-8") as f:
            data = json.load(f)
        cards = data.get("datasets", [])
        errors = data.get("errors", [])
        print(f"Loaded {len(cards)} cards + {len(errors)} errors from {src}")
        _generate_figures(cards, errors, png_dir, pdf_path)
        return

    # ── Mode: scan datasets ───────────────────────────────────────────
    if not DATA_ROOT.exists():
        print(f"Data root not found: {DATA_ROOT}")
        sys.exit(1)

    # Load selection filter from DatabaseDetail.xlsx
    selected_suffixes = None
    if args.all:
        print("--all: scanning all datasets (ignoring selection file)")
    elif SELECTION_XLSX.exists():
        selected_suffixes = load_selected_suffixes(SELECTION_XLSX)
        print(f"Selection filter: {len(selected_suffixes)} datasets from {SELECTION_XLSX.name}")
    else:
        print(f"WARNING: {SELECTION_XLSX.name} not found — scanning all datasets")

    # Determine scan root(s)
    if args.category:
        scan_root = DATA_ROOT / args.category
        if not scan_root.exists():
            print(f"Category folder not found: {scan_root}")
            sys.exit(1)
        roots = [scan_root]
    else:
        roots = []
        for sub in sorted(DATA_ROOT.iterdir()):
            if sub.is_dir():
                roots.append(sub)

    # Find all datasets (filtered by selection)
    all_datasets = []
    for root in roots:
        all_datasets.extend(find_datasets(root, selected_suffixes))

    print(f"\nFound {len(all_datasets)} datasets\n")

    cards = []
    errors = []

    for i, info in enumerate(all_datasets):
        print(f"[{i+1}/{len(all_datasets)}] Loading {info['name']}...", end=" ", flush=True)
        try:
            card = build_id_card(info)
            cards.append(card)
            print("OK")
            print_card(card)
            print()
        except Exception as e:
            error_msg = f"{type(e).__name__}: {e}"
            tb = traceback.format_exc()
            errors.append({
                "name": info["name"],
                "path": info["path"],
                "groups": info["groups"],
                "error": error_msg,
                "traceback": tb,
            })
            print("FAILED")
            print_error(info, error_msg)
            print()

    # Summary
    print("\n" + "=" * 72)
    print(f"  SUMMARY: {len(cards)} loaded, {len(errors)} errors")
    print("=" * 72)

    if errors:
        print("\n  Datasets with errors:")
        for err in errors:
            print(f"    - {err['name']}: {err['error']}")

    # JSON output
    output = {
        "datasets": cards,
        "errors": errors,
        "summary": {
            "total_found": len(all_datasets),
            "loaded": len(cards),
            "failed": len(errors),
        }
    }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False, default=str)
    print(f"\nJSON written to: {json_path}")

    # Figure generation
    if not args.no_plots:
        _generate_figures(cards, errors, png_dir, pdf_path)

def _generate_figures(cards: list[dict], errors: list[dict], png_dir: Path, pdf_path: Path):
    """Generate PNG per dataset + combined PDF."""
    png_dir.mkdir(exist_ok=True)
    total = len(cards) + len(errors)

    print(f"\nGenerating {total} figures into {png_dir}/")

    for i, card in enumerate(cards):
        png_path = png_dir / f"{card['name']}.png"
        print(f"  [{i+1}/{total}] {card['name']}.png", end=" ", flush=True)
        try:
            render_card_figure(card, png_path)
            print("OK")
        except Exception as e:
            print(f"FAILED: {e}")

    for j, err in enumerate(errors):
        png_path = png_dir / f"{err['name']}_error.png"
        print(f"  [{len(cards)+j+1}/{total}] {err['name']}_error.png", end=" ", flush=True)
        try:
            render_error_figure(err, png_path)
            print("OK")
        except Exception as e:
            print(f"FAILED: {e}")

    # Combined PDF
    print(f"\nBuilding PDF: {pdf_path} ...", end=" ", flush=True)
    render_pdf(cards, errors, pdf_path, png_dir)
    print("OK")

if __name__ == "__main__":
    main()
